#!/usr/bin/env python
# -*- coding: utf-8 -*-

########################
#                      #
#        PASOS         #
#                      #
########################

# Unidades
# Outliers
# Factor rinde real
# Armado de grilla
# Interseccion de mapa de ambientes con mapa de parcelas
# Interseccion de grilla con mapa de ambientes_parcelas
# Calcular superficie de cada poligono
# Interseccion con el mapa de puntos
# Listado de todos los puntos de los cuadrados enteros por parcela y ambiente
# Corroborar si existe la misma cantidad de ambientes por parcela > errs
# Descartar los cuadrados con muchos o pocos puntos (threshold)
# Seleccionar cantidad minima de puntos por cuadrado al azar
# Seleccionar cantidad minima de cuadrados por ambiente al azar
## Exportar shp de puntos seleccionados ?
# Exportar excel de puntos seleccionados
# Supuestos
# ANOVA
# Comparacion de medias

import arcpy
from arcpy import env
import numpy as np
import copy
import random
import openpyxl
import pandas as pd
from scipy import stats
import math
from statsmodels.formula.api import ols
import matplotlib.pyplot as plt
import os
import time

#
# PARAMS
# 

"""
out_dir = r"E:\ArcGIS\test2"# arcpy.GetParameterAsText(0) #r"E:\ArcGIS\test"
parcelas_shp = out_dir + r"\Export_Output_2.shp"# arcpy.GetParameterAsText(1) #out_dir + r"\parcelas.shp"
ambientes_shp = out_dir + r"\Export_Output_5.shp"# arcpy.GetParameterAsText(2) #out_dir + r"\ambientes_2.shp"
puntos_shp = out_dir + r"\Export_Output_4.shp"# arcpy.GetParameterAsText(3) #out_dir + r"\SG1_Merge.shp"
lotes_shp = out_dir + r"\Export_Output.shp"# arcpy.GetParameterAsText(4) #out_dir + r"\lotes.shp"
rindes_xlsx = out_dir +  r"\rindes.xlsx"# arcpy.GetParameterAsText(5) #out_dir +  r"\rindes.xlsx"
cell_size = 21# arcpy.GetParameterAsText(6) #50
#cell_num = # arcpy.GetParameterAsText(7) #10
todos_los_ambientes = False# arcpy.GetParameterAsText(7) #False
campo_idAlbor = "idAlbor"# arcpy.GetParameterAsText(8) #"idAlbor"
campo_rendimiento = "Yld_Mass_W"# arcpy.GetParameterAsText(9) #"Rendimient"
campo_parcela = "Parcela"# arcpy.GetParameterAsText(10) #"Parcela"
campo_ambiente = "Ambiente"# arcpy.GetParameterAsText(11) #"Ambiente"
"""

out_dir = arcpy.GetParameterAsText(0) #r"E:\ArcGIS\test"
parcelas_shp = arcpy.GetParameterAsText(1) #out_dir + r"\parcelas.shp"
ambientes_shp = arcpy.GetParameterAsText(2) #out_dir + r"\ambientes_2.shp"
puntos_shp = arcpy.GetParameterAsText(3) #out_dir + r"\SG1_Merge.shp"
lotes_shp = arcpy.GetParameterAsText(4) #out_dir + r"\lotes.shp"
rep_config = arcpy.GetParameterAsText(5) #20
amend_yield = arcpy.GetParameterAsText(6) #True
rindes_xlsx = arcpy.GetParameterAsText(7) #out_dir +  r"\rindes.xlsx"
use_fishnet = arcpy.GetParameterAsText(8) #True
cell_size = arcpy.GetParameterAsText(9) #50
buffer = arcpy.GetParameterAsText(10) #0
todos_los_ambientes = arcpy.GetParameterAsText(11) #False
campo_idAlbor = arcpy.GetParameterAsText(12) #"idAlbor"
campo_campo = arcpy.GetParameterAsText(13) #"idAlbor"
campo_rendimiento = arcpy.GetParameterAsText(14) #"Rendimient"
campo_parcela = arcpy.GetParameterAsText(15) #"Parcela"
amend_parcelas = arcpy.GetParameterAsText(16) #True
campo_ambiente = arcpy.GetParameterAsText(17) #"Ambiente"
nombre_analisis = arcpy.GetParameterAsText(18) #"Test"
exportar_aux = arcpy.GetParameterAsText(19) #False
exportar_intersect = arcpy.GetParameterAsText(20) #True
exportar_mapa_rinde = arcpy.GetParameterAsText(21) #True
dca1f_por_ambiente = arcpy.GetParameterAsText(22) #True
perdonar_falta_cuadrados = arcpy.GetParameterAsText(23) #False
nombre_testigo = arcpy.GetParameterAsText(24) #"Testigo"

#
# CONFIG
# 

arcpy.env.overwriteOutput = True
env.outputCoordinateSystem = arcpy.SpatialReference("WGS 1984 UTM Zone 21S")

timestamp = str(int(time.time()*1000))
out_data_xlsx = out_dir +  r"\data_" + timestamp + ".xlsx"
aditividad = out_dir +  r"\aditividad.png"
resvspred = out_dir +  r"\res_vs_pred.png"
qqplot = out_dir +  r"\qq_plot.png"

arcpy.Delete_management("in_memory")
out_temp = "in_memory"

distancia = 20 # Distancia eucladiana moran I
gl_obj =  40 # Grados de libertad (asegura 40 gl ajustando la cantidad de repeticiones, si la cantidad de repeticiones en la config es mayor, toma ese valor)
tipo = "dca1f" if use_fishnet == True or use_fishnet == "true" else "dca2f" # "dca2f" || "dbca1f" || "dca1f"
puntos_min_abs_config = 10 
puntos_min_config = max(puntos_min_abs_config, int(int(cell_size) * 0.8)) # El tamaño del cuadrado deberia ser igual para todos los ensayos

########################
#                      #
#   PREPROCESAMIENTO   #
#                      #
########################

"""
{
    "Testigo": {
        "A": {
            "1": [...],
            "2": [...]
        },
        "B": {...},
        "C": {...}
    },
    "Ensayo": {...},
}
"""

new_out_intersect_3 = out_dir + r"\mapa_rinde.shp"
out_intersect_3 = out_temp + r"\intersect_3"

if os.path.exists(new_out_intersect_3) == True:
    arcpy.CopyFeatures_management(new_out_intersect_3, out_intersect_3)
    campo_rendimiento = "src_Rinde"
    ID_MAPA_RINDE = "ID_MAPA_RI"
else:
    ID_MAPA_RINDE = "ID_MAPA_RINDE"

    #
    # UNIDADES
    #

    mem_puntos_shp = out_temp + r"\puntos_shp"
    arcpy.CopyFeatures_management(puntos_shp,mem_puntos_shp)

    i = 0
    suma = 0
    for row in arcpy.SearchCursor(mem_puntos_shp):
        i += 1
        suma += row.getValue(campo_rendimiento)
        if i == 500:
            break

    if not "src_Rinde" in [field.name for field in arcpy.ListFields(mem_puntos_shp)]:
        arcpy.AddField_management(mem_puntos_shp, "src_Rinde", "DOUBLE")

    if suma/i < 1000:
        arcpy.AddMessage("Rinde en toneladas")
        arcpy.CalculateField_management(mem_puntos_shp, "src_Rinde","!" + campo_rendimiento + "! * 1000","PYTHON")
    else:
        arcpy.CalculateField_management(mem_puntos_shp, "src_Rinde","!" + campo_rendimiento + "!","PYTHON")
        arcpy.AddMessage("Rinde en kilos")

    campo_rendimiento = "src_Rinde"
    arcpy.AddMessage("")

    #
    # OUTLIERS
    #

    out_zero = out_temp + r"\zero"
    print "Filtering Min-Max Outliers"
    arcpy.Select_analysis (mem_puntos_shp, out_zero, '"' + campo_rendimiento + '" > 0')

    out_min_max = out_temp + r"\min_max"
    print "Filtering Min-Max Outliers"
    puntos = [row.getValue(campo_rendimiento) for row in arcpy.SearchCursor(out_zero)]
    rinde_max = np.percentile(puntos, 75) + 1.5 * (np.percentile(puntos, 75) - np.percentile(puntos, 25))
    rinde_min = np.percentile(puntos, 25) - 1.5 * (np.percentile(puntos, 75) - np.percentile(puntos, 25))
    arcpy.Select_analysis (out_zero, out_min_max, '"' + campo_rendimiento + '" < ' + str(rinde_max) + ' AND "' + campo_rendimiento + '" > ' + str(rinde_min))

    out_outliers = out_temp + r"\outliers"
    print "Finding Positional Outliers"
    arcpy.ClustersOutliers_stats(out_min_max, campo_rendimiento, out_outliers, "INVERSE_DISTANCE", "EUCLIDEAN_DISTANCE", "NONE", distancia)

    out_filtered = out_temp + r"\filtered"
    print "Filtering Outliers"
    arcpy.Select_analysis (out_outliers, out_filtered, "\"COType\" = ' ' OR \"COType\" = 'HH' OR \"COType\" = 'LL'")
    arcpy.AddField_management(out_filtered, ID_MAPA_RINDE, "LONG")
    arcpy.CalculateField_management(out_filtered, ID_MAPA_RINDE, "!SOURCE_ID!","PYTHON")

    print ""

    #
    # FACTOR RINDE REAL
    # 

    arcpy.Intersect_analysis([lotes_shp, out_filtered], out_intersect_3)

    if amend_yield == True or amend_yield == "true":
        lotes = set( [ row.getValue(campo_idAlbor) for row in arcpy.SearchCursor(lotes_shp) ] )

        wb = openpyxl.load_workbook(rindes_xlsx,data_only=True)
        ws = wb['Rinde']

        # arcpy.Intersect_analysis([lotes_shp, out_filtered], out_intersect_3)

        for idAlbor in lotes:
            rinde_mapa = np.mean( [ row.getValue(campo_rendimiento) for row in arcpy.SearchCursor(out_intersect_3) if row.getValue(campo_idAlbor) == idAlbor ] )

            rinde_real = 0
            for row in ws.values:
                if row[0] == "":
                    break
                if row[0] == idAlbor:
                    rinde_real = row[2]
                    break
            
            if rinde_real > 0:
                factor = 1 + ((rinde_real - rinde_mapa) / rinde_mapa)
                mensaje = "idAlbor: " + str(idAlbor) + ", Rinde mapa: " + str(round(rinde_mapa,2)) + ", Rinde real: " + str(round(rinde_real,2)) + ", Factor: " + str(factor)

                code_block = """
def NewVal(campo_idAlbor,old_val):
    if campo_idAlbor == """ + str(idAlbor) + """:
        return old_val * """ + str(factor) + """
    else:
        return old_val"""

                arcpy.CalculateField_management(out_intersect_3, campo_rendimiento, "NewVal(!" + campo_idAlbor + "!, !" + campo_rendimiento + "!)", "PYTHON", code_block)
            else:
                raise Exception("Rinde real no encontrado en excel")

            arcpy.AddMessage(mensaje)
            arcpy.AddMessage("")
    """
    else:
        out_intersect_3 = out_filtered
    """
    
    if exportar_mapa_rinde == True or exportar_mapa_rinde == "true":
        arcpy.CopyFeatures_management(out_intersect_3, new_out_intersect_3)

#
# CAMPO
# 

nombre_campo = ''
for row in arcpy.SearchCursor(out_intersect_3):
    nombre_campo = row.getValue(campo_campo)
    break

#
# LIMPIA CAMPOS
# 

fields = [field.name for field in arcpy.ListFields(out_intersect_3) if field.name != campo_rendimiento and field.name != ID_MAPA_RINDE and field.name != "FID" and field.name != "Shape"]
if len(fields) > 0:
    arcpy.DeleteField_management(out_intersect_3, fields)

#
# FISHNET
# 

if use_fishnet == True or use_fishnet == "true":
    extent = arcpy.Describe(ambientes_shp).extent
    origin_coord = " ".join([str(extent.XMin),str(extent.YMin)])
    y_axis_coord = " ".join([str(extent.XMin),str(extent.YMax)])
    corner_coord = " ".join([str(extent.XMax),str(extent.YMax)])

    out_fishnet = out_temp + r"\fishnet"
    arcpy.CreateFishnet_management(out_fishnet, origin_coord, y_axis_coord, cell_size, cell_size, 0, 0, corner_coord, "NO_LABELS", "#", "POLYGON")

    if exportar_aux == True or exportar_aux == "true":
        new_out_fishnet = out_dir + r"\fishnet.shp"
        arcpy.CopyFeatures_management(out_fishnet,new_out_fishnet)

#
# PARCELAS SHP
# 

mem_parcelas_shp = out_temp + r"\mem_parcelas_shp"
arcpy.CopyFeatures_management(parcelas_shp, mem_parcelas_shp)

if not campo_parcela in [field.name for field in arcpy.ListFields(mem_parcelas_shp, field_type='String')]:
    if not "str_Parcela" in [field.name for field in arcpy.ListFields(mem_parcelas_shp)]:
        arcpy.AddField_management(mem_parcelas_shp, "str_Parcela", "TEXT")
    arcpy.CalculateField_management(mem_parcelas_shp, "str_Parcela", "str(!" + campo_parcela + "!)", "PYTHON")

    campo_parcela = "str_Parcela"

if amend_parcelas == True or amend_parcelas == "true":
    if not "src_Parcela" in [field.name for field in arcpy.ListFields(mem_parcelas_shp)]:
        arcpy.AddField_management(mem_parcelas_shp, "src_Parcela", "TEXT")

    code_block = """
def NewVal(campo_parcela):
    if campo_parcela == '0':
        return 'Testigo'
    else:
        return 'Ensayo'"""

    arcpy.CalculateField_management(mem_parcelas_shp, "src_Parcela", "NewVal(!" + campo_parcela + "!)", "PYTHON", code_block)

    campo_parcela = "src_Parcela"
    nombre_testigo = "Testigo"


#
# LIMPIA CAMPOS
# 

fields = [field.name for field in arcpy.ListFields(mem_parcelas_shp) if field.name != campo_parcela and field.name != "FID" and field.name != "Shape"]
if len(fields) > 0:
    arcpy.DeleteField_management(mem_parcelas_shp, fields)

#
# AMBIENTES SHP
# 

mem_ambientes_shp = out_temp + r"\mem_ambientes_shp"
arcpy.CopyFeatures_management(ambientes_shp, mem_ambientes_shp)

if not campo_ambiente in [field.name for field in arcpy.ListFields(mem_ambientes_shp, field_type='String')]:
    if not "str_Ambiente" in [field.name for field in arcpy.ListFields(mem_ambientes_shp)]:
        arcpy.AddField_management(mem_ambientes_shp, "str_Ambiente", "TEXT")
    arcpy.CalculateField_management(mem_ambientes_shp, "str_Ambiente", "str(!" + campo_ambiente + "!)", "PYTHON")

    campo_ambiente = "str_Ambiente"

#
# LIMPIA CAMPOS
# 

fields = [field.name for field in arcpy.ListFields(mem_ambientes_shp) if field.name != campo_ambiente and field.name != "FID" and field.name != "Shape"]
if len(fields) > 0:
    arcpy.DeleteField_management(mem_ambientes_shp, fields)

#
# INTERSECT
# 

out_intersect_0 = out_temp + r"\intersect_0"
arcpy.Intersect_analysis([mem_ambientes_shp, mem_parcelas_shp], out_intersect_0)

#
# BUFFER
# 

if int(buffer) > 0:
    out_buffer = out_temp + r"\buffer"
    arcpy.Buffer_analysis(out_intersect_0, out_buffer, "-" + buffer + " Meters")
else:
    if not (use_fishnet == True or use_fishnet == "true"):
        raise Exception("Ingrese un valor para el buffer mayor a 0!")
    out_buffer = out_intersect_0

#
# INTERSECT
# 

if use_fishnet == True or use_fishnet == "true":
    out_intersect_1 = out_temp + r"\intersect_1"
    arcpy.Intersect_analysis([out_buffer, out_fishnet], out_intersect_1)
else:
    out_intersect_1 = out_buffer

#
# AREA
# 

if not "src_AREA" in [field.name for field in arcpy.ListFields(out_intersect_1)]:
    arcpy.AddField_management(out_intersect_1, "src_AREA", "DOUBLE")

arcpy.CalculateField_management(out_intersect_1, "src_AREA","!shape.area@hectares!!","PYTHON_9.3")

#
# INTERSECT
# 

out_intersect_2 = out_temp + r"\intersect_2"
arcpy.Intersect_analysis([out_intersect_3, out_intersect_1], out_intersect_2)

if exportar_aux == True or exportar_aux == "true":
    new_out_fishnet = out_dir + r"\intersect_2.shp"
    arcpy.CopyFeatures_management(out_intersect_2,new_out_fishnet)

#
# LIST
# 

cell_area = round(float(cell_size) * float(cell_size) / 10000, 2) if use_fishnet == True or use_fishnet == "true" else 0

cells_dict = {}
for row in arcpy.SearchCursor(out_intersect_2):
    area = row.getValue("src_AREA")
    parcela = row.getValue(campo_parcela)
    ambiente = row.getValue(campo_ambiente)
    rinde = row.getValue(campo_rendimiento)
    fid = row.getValue(ID_MAPA_RINDE)
    #fid = row.getValue("ID_MAPA_RI")

    if round(area, 2) < cell_area:
        continue
    
    if parcela not in cells_dict:
        cells_dict[parcela] = {}

    if ambiente not in cells_dict[parcela]:
        cells_dict[parcela][ambiente] = {}

    if use_fishnet == True or use_fishnet == "true":
        fishnet_id = row.getValue("FID_fishnet")
    else:
        fishnet_id = max(len(cells_dict[parcela][ambiente].keys()) - 1, 0)

    if fishnet_id not in cells_dict[parcela][ambiente]:
        cells_dict[parcela][ambiente][fishnet_id] = []
    elif len(cells_dict[parcela][ambiente][fishnet_id]) >= puntos_min_abs_config + 1 and (use_fishnet == False or use_fishnet == "false"):
        fishnet_id += 1
        cells_dict[parcela][ambiente][fishnet_id] = []
    
    cells_dict[parcela][ambiente][fishnet_id].append([rinde,fid])
    #cells_dict[parcela][ambiente][fishnet_id].append(rinde)

#
# PARCELAS
#

parcelas = len(cells_dict.keys())
arcpy.AddMessage("Parcelas: " + str(parcelas))

if len(set( [ row.getValue(campo_parcela) for row in arcpy.SearchCursor(mem_parcelas_shp) ] )) != parcelas:
    raise Exception("La cantidad de parcelas del muestreo no coincide con las parcelas del shp de parcelas!")

#
# AMBIENTES x PARCELA
#

if todos_los_ambientes == False or todos_los_ambientes == "false":

    ambientes_dict = {}
    for parcela in cells_dict:
        for ambiente in cells_dict[parcela]:
            
            if ambiente not in ambientes_dict:
                ambientes_dict[ambiente] = 0

            ambientes_dict[ambiente] += 1
       
    ambientes_dict_c = copy.deepcopy(ambientes_dict)
    for ambiente in ambientes_dict_c:
        if ambientes_dict[ambiente] < parcelas:
            del ambientes_dict[ambiente]
    
    ambientes = len(ambientes_dict.keys())
    if ambientes == 0:
        raise Exception("Hay parcelas que quedan sin ambientes!")
    
    # Limpia el dict de los ambientes que no tienen todas las parcelas
    cells_dict_c = copy.deepcopy(cells_dict)
    for parcela in cells_dict_c:
        for ambiente in cells_dict_c[parcela]:
            if ambiente not in ambientes_dict:
                del cells_dict[parcela][ambiente]
    
else:
    ambientes = len(set( [ row.getValue(campo_ambiente) for row in arcpy.SearchCursor(ambientes_shp) ] ))

    for parcela in cells_dict:
        if len(cells_dict[parcela].keys()) != ambientes:
            raise Exception("La cantidad de ambientes por parcela del muestreo no coincide con la cantidad de ambientes del shp de ambientes!")

if ambientes == 1 or dca1f_por_ambiente == True or dca1f_por_ambiente == "true":
    tipo = "dca1f"

arcpy.AddMessage("Ambientes por parcela: " + str(ambientes))

if tipo == "dca1f":
    min_rep = math.ceil(gl_obj / parcelas + 1)
elif tipo == "dbca1f":
    min_rep = math.ceil(gl_obj / parcelas)
elif tipo == "dca2f":
    min_rep = math.ceil(gl_obj / (parcelas * ambientes) + 1)

min_rep = max(int(rep_config), int(min_rep))

"""
#
# PUNTOS X CUADRADO (limpieza de cuadrados con menos puntos que el minimo)
# 

cells_dict_c = copy.deepcopy(cells_dict)
for parcela in cells_dict_c:
    for ambiente in cells_dict_c[parcela]:
        for cuadrado in cells_dict_c[parcela][ambiente]:

            if len(cells_dict[parcela][ambiente][cuadrado]) < puntos_min:
                del cells_dict[parcela][ambiente][cuadrado]
"""

#
# PUNTOS X CUADRADO (limpieza de cuadrados con pocos o muchos puntos)
# 

puntos = []
for parcela in cells_dict:
    for ambiente in cells_dict[parcela]:
        for cuadrado in cells_dict[parcela][ambiente]:
            puntos.append(len(cells_dict[parcela][ambiente][cuadrado]))

puntos_max = np.min([ np.percentile(puntos, 75) + 1.5 * (np.percentile(puntos, 75) - np.percentile(puntos, 25)), np.max(puntos) ])
puntos_min = np.max([ np.percentile(puntos, 25) - 1.5 * (np.percentile(puntos, 75) - np.percentile(puntos, 25)), puntos_min_config + 1 ])

if puntos_min > puntos_max:
    puntos_min = puntos_min_abs_config + 1

if puntos_min > puntos_max:
    raise Exception("La cantidad de puntos por cuadrado no es suficiente! " + str(puntos_max)  + "/" + str(puntos_min))

cells_dict_c = copy.deepcopy(cells_dict)
for parcela in cells_dict_c:
    for ambiente in cells_dict_c[parcela]:
        for cuadrado in cells_dict_c[parcela][ambiente]:
            if len(cells_dict[parcela][ambiente][cuadrado]) > puntos_max or len(cells_dict[parcela][ambiente][cuadrado]) < puntos_min:
                del cells_dict[parcela][ambiente][cuadrado]

#
# CUADRADO X AMBIENTE (seleccion al azar de cantidad minima de cuadrados por ambiente)
# 

aux_forgiveness = pd.DataFrame(columns=["Parcelas","Diferencia_estadistica","Rinde_Parcela","Rinde_Testigo","Nombre","Campo","Ambiente","DMS","Diferencia"])

if perdonar_falta_cuadrados == True or perdonar_falta_cuadrados == "true":
    ambientes_flojos = set()
    for parcela in cells_dict:
        for ambiente in cells_dict[parcela]:
            if len(cells_dict[parcela][ambiente]) < min_rep + 1:
                ambientes_flojos.add(ambiente)

    rinde_dict = {}
    for parcela in cells_dict_c:
        rinde_dict[parcela] = {}
        for ambiente in cells_dict_c[parcela]:
            if ambiente in ambientes_flojos:
                
                rindes = []
                for cuadrado in cells_dict[parcela][ambiente]:
                    rindes += [ p[0] for p in cells_dict[parcela][ambiente][cuadrado] ]
                rinde = np.median(rindes)
                
                rinde_dict[parcela][ambiente] = rinde

    cells_dict_c = copy.deepcopy(cells_dict)
    for parcela in cells_dict_c:
        for ambiente in cells_dict_c[parcela]:
            if ambiente in ambientes_flojos:

                if parcela != nombre_testigo:
                    aux_forgiveness.loc[len(aux_forgiveness)] = {
                        "Parcelas": parcela,
                        "Diferencia_estadistica": 0,
                        "Rinde_Parcela": rinde_dict[parcela][ambiente],
                        "Rinde_Testigo": rinde_dict[nombre_testigo][ambiente],
                        "Nombre": nombre_analisis,
                        "Campo": nombre_campo,
                        "Ambiente": ambiente,
                        "DMS": 0,
                        "Diferencia": rinde_dict[parcela][ambiente] - rinde_dict[nombre_testigo][ambiente]
                    }
                
                del cells_dict[parcela][ambiente]
    
    for ambiente in ambientes_flojos:
        arcpy.AddMessage("Perdonado: " + ambiente)

cuadrados = []
for parcela in cells_dict:
    for ambiente in cells_dict[parcela]:
        cuadrados.append(len(cells_dict[parcela][ambiente]))

#cuadrados_max = np.percentile(cuadrados, 75) + 1.5 * (np.percentile(cuadrados, 75) - np.percentile(cuadrados, 25))
#cuadrados_min = np.percentile(cuadrados, 25) - 1.5 * (np.percentile(cuadrados, 75) - np.percentile(cuadrados, 25))
#print np.percentile(cuadrados, 50),cuadrados_min,cuadrados_max

cuadrados_min = np.min(cuadrados)

if cuadrados_min < min_rep + 1:
    raise Exception("La cantidad de cuadrados por ambiente no es suficiente! " + str(cuadrados_min) + "/" + str(min_rep + 1))

arcpy.AddMessage("Cuadrados por ambiente (r): " + str(min_rep))

cells_dict_c = copy.deepcopy(cells_dict)
for parcela in cells_dict:
    for ambiente in cells_dict[parcela]:
        cuadrados_rnd = random.sample(cells_dict[parcela][ambiente], min_rep)
        cells_dict[parcela][ambiente] = {}

        for cuadrado in cuadrados_rnd:
            cells_dict[parcela][ambiente][cuadrado] = cells_dict_c[parcela][ambiente][cuadrado]

#
# PUNTOS X CUADRADO (seleccion al azar de cantidad minima de puntos por cuadrado)
# 

puntos = []
for parcela in cells_dict:
    for ambiente in cells_dict[parcela]:
        for cuadrado in cells_dict[parcela][ambiente]:
            puntos.append(len(cells_dict[parcela][ambiente][cuadrado]))

puntos_min = np.min(puntos) - 1

ids = []
for parcela in cells_dict:
    for ambiente in cells_dict[parcela]:
        for cuadrado in cells_dict[parcela][ambiente]:
            puntos = random.sample(cells_dict[parcela][ambiente][cuadrado], puntos_min)
            rindes = [ p[0] for p in puntos ]
            ids += [ p[1] for p in puntos ]
            cells_dict[parcela][ambiente][cuadrado] = [np.median(rindes)]

arcpy.AddMessage("Puntos por cuadrado: 1 (" + str(puntos_min) + ")")
arcpy.AddMessage("")

#
# EXPORTAR EXCEL
# 

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "DATA"

ws.cell(row=1, column=1, value="Parcela")
ws.cell(row=1, column=2, value="Ambiente")
ws.cell(row=1, column=3, value="ID_fishnet")
ws.cell(row=1, column=4, value="Rinde")

r = 2
data = {
    "Parcela": [],
    "Ambiente": [],
    "ID_fishnet": [],
    "Rinde":[]
}

for parcela in cells_dict:
    for ambiente in cells_dict[parcela]:
        for cuadrado in cells_dict[parcela][ambiente]:
            for rinde in cells_dict[parcela][ambiente][cuadrado]: #Hay uno solo por cuadrado (mediana)

                ws.cell(row=r, column=1, value=parcela)
                data["Parcela"].append(parcela)
                ws.cell(row=r, column=2, value=ambiente)
                data["Ambiente"].append(ambiente)
                ws.cell(row=r, column=3, value=cuadrado)
                data["ID_fishnet"].append(cuadrado)
                ws.cell(row=r, column=4, value=rinde)
                data["Rinde"].append(rinde)

                r += 1

if exportar_intersect == True or exportar_intersect == "true":
    new_out_intersect = out_dir + r"\intersect.shp"
    #arcpy.Select_analysis(out_intersect_2, new_out_intersect, "\"FID_fishnet\" IN (" + ",".join([str(i) for i in data["ID_fishnet"] ])  + ")") 
    arcpy.Select_analysis(out_intersect_2, new_out_intersect, "\"" + ID_MAPA_RINDE + "\" IN (" + ",".join([str(i) for i in ids ])  + ")") 
    arcpy.AddMessage("Intersect exportado")

data = pd.DataFrame(data)
wb.save(out_data_xlsx)
wb.close()

arcpy.AddMessage("Excel exportado")
arcpy.AddMessage("")
arcpy.AddMessage("")


########################
#                      #
# ANALISIS ESTADISTICO #
#                      #
########################

def DCA1F(data, columnas):
    """
    Diseño completo al azar

    Columnas: [y, x]
    """

    N = len(data[columnas[0]])
    df_a = len(data[columnas[1]].unique()) - 1
    df_e = N - len(data[columnas[1]].unique())
    df_t = N - 1

    print N,len(data[columnas[1]].unique())

    grand_mean = data[columnas[0]].mean()
    ssq_a = sum([(data[data[columnas[1]] == i][columnas[0]].mean() - grand_mean) ** 2 for i in data[columnas[1]]])
    ssq_t = sum((data[columnas[0]] - grand_mean) ** 2)
    ssq_e = ssq_t - ssq_a

    ms_a = ssq_a/df_a
    ms_e = ssq_e/df_e

    f_a = ms_a/ms_e

    p_a = stats.f.sf(f_a, df_a, df_e)

    results = {
        "SC": [ssq_a, ssq_e, ssq_t],
        "gl": [df_a, df_e, df_t],
        "CM": [ms_a, ms_e, ""],
        "F": [f_a, "", ""],
        "p-valor": [p_a, "", ""]
    }

    columns = ["SC", "gl", "CM", "F", "p-valor"]

    return pd.DataFrame(results, columns=columns, index=[columnas[1] + "s", "Error", "Total"])

def DBCA1F(data, columna):
    """
    Diseño en bloques completos al azar

    Los bloques son las repeticiones
    """

    df_a = len(data.Parcela.unique()) - 1
    df_b = len(data.Ambiente.unique()) - 1
    df_e = df_a * df_b 
    df_t = len(data.Parcela.unique()) * len(data.Ambiente.unique()) - 1

    grand_mean = data[columna].mean()
    ssq_a = sum([(data[data.Parcela == i][columna].mean() - grand_mean) ** 2 for i in data.Parcela])
    ssq_b = sum([(data[data.Ambiente == i][columna].mean() - grand_mean) ** 2 for i in data.Ambiente])
    ssq_t = sum((data[columna] - grand_mean) ** 2)
    ssq_e = ssq_t - ssq_a - ssq_b

    ms_a = ssq_a/df_a
    ms_b = ssq_b/df_b
    ms_e = ssq_e/df_e

    f_a = ms_a/ms_e
    f_b = ms_b/ms_e

    p_a = stats.f.sf(f_a, df_a, df_e)
    p_b = stats.f.sf(f_b, df_b, df_e)

    results = {
        "SC": [ssq_a, ssq_b, ssq_e, ssq_t],
        "gl": [df_a, df_b, df_e, df_t],
        "CM": [ms_a, ms_b, ms_e, ""],
        "F": [f_a, f_b, "", ""],
        "p-valor": [p_a, p_b, "", ""]
    }

    columns = ["SC", "gl", "CM", "F", "p-valor"]

    return pd.DataFrame(results, columns=columns, index=["Parcelas", "Ambientes", "Error", "Total"])

def DCA2F(data, columna):
    """
    Diseño completo al azar con dos factores
    """
    
    N = len(data[columna])
    df_a = len(data.Parcela.unique()) - 1
    df_b = len(data.Ambiente.unique()) - 1
    df_axb = df_a * df_b
    df_e = N - (len(data.Parcela.unique()) * len(data.Ambiente.unique()))
    df_t = N - 1

    grand_mean = data[columna].mean()
    ssq_a = sum([(data[data.Parcela == i][columna].mean() - grand_mean) ** 2 for i in data.Parcela])
    ssq_b = sum([(data[data.Ambiente == i][columna].mean() - grand_mean) ** 2 for i in data.Ambiente])
    ssq_t = sum((data[columna] - grand_mean) ** 2)

    ssq_e = 0
    for ambiente in data.Ambiente.unique():
        sub_data = data[data.Ambiente == ambiente]
        means = [sub_data[sub_data.Parcela == i][columna].mean() for i in sub_data.Parcela]
        ssq_e += sum((sub_data[columna] - means) ** 2)

    ssq_axb = ssq_t - ssq_a - ssq_b - ssq_e

    ms_a = ssq_a/df_a
    ms_b = ssq_b/df_b
    ms_axb = ssq_axb/df_axb
    ms_e = ssq_e/df_e

    f_a = ms_a/ms_e
    f_b = ms_b/ms_e
    f_axb = ms_axb/ms_e

    p_a = stats.f.sf(f_a, df_a, df_e)
    p_b = stats.f.sf(f_b, df_b, df_e)
    p_axb = stats.f.sf(f_axb, df_axb, df_e)

    results = {
        "SC": [ssq_a, ssq_b, ssq_axb, ssq_e, ssq_t],
        "gl": [df_a, df_b, df_axb, df_e, df_t],
        "CM": [ms_a, ms_b, ms_axb, ms_e, ""],
        "F": [f_a, f_b, f_axb, "", ""],
        "p-valor": [p_a, p_b, p_axb, "", ""]
    }

    columns = ["SC", "gl", "CM", "F", "p-valor"]

    return pd.DataFrame(results, columns=columns, index=["Parcelas", "Ambientes", "Parcelas * Ambientes", "Error", "Total"])

def Assumptions(data,writer,ambiente):
    """
    Normalidad
        Residuos absolutos
        Si la cantidad de datos > 50 >> Kolmogorov-Smirnov
        Si la cantidad de datos < 50 >> Shapiro-Wilks
        QQ-plot

    Homocedasticidad
        Levene: ANOVA usando residuos absolutos
        Res vs Pred: Residuos estudentizados versus valores predichos

    Aditividad
        Para bloques o factores
        Promedio de rinde para cada ambiente por tratamiento

    """

    arcpy.AddMessage("SUPUESTOS")
    arcpy.AddMessage("")

    if tipo == "dca1f":
        model = ols("Rinde ~ C(Parcela)", data=data).fit()
    elif tipo == "dbca1f":
        model = ols("Rinde ~ C(Ambiente) + C(Parcela)", data=data).fit() 
    elif tipo == "dca2f":
        model = ols("Rinde ~ C(Ambiente) + C(Parcela) + C(Parcela):C(Ambiente)", data=data).fit() 

    #arcpy.AddMessage(model.summary())
    #arcpy.AddMessage("")

    #
    # NORMALIDAD
    #

    arcpy.AddMessage("Normalidad")
    if len(data.Rinde) < 50:
        _, pvalue = stats.shapiro(model.resid)
        pd.DataFrame({"Shapiro-Wilks": [pvalue]}).to_excel(writer, "NORMALIDAD" + ambiente, index=False)
        arcpy.AddMessage(" ".join(["Shapiro-Wilks:",str(pvalue)]))
    else:
        _, pvalue = stats.kstest(model.resid, "norm")
        pd.DataFrame({"Kolmogorov-Smirnov": [pvalue]}).to_excel(writer, "NORMALIDAD" + ambiente, index=False)
        arcpy.AddMessage(" ".join(["Kolmogorov-Smirnov:",str(pvalue)]))

    """
    AGREGAR + ambiente
    #stats.probplot(model.resid, dist="norm", plot=plt)
    #arcpy.AddMessage("Ver gráfico")
    #plt.show()
    #plt.savefig(qqplot)
    """

    arcpy.AddMessage("")

    #
    # HOMOCEDASTICIDAD
    #

    data["Residuos_ABS"] = [abs(i) for i in model.resid]
    data["Residuos_Est"] = model.outlier_test()["student_resid"]
    data["Predichos"] = model.fittedvalues

    if tipo == "dca1f":
        aov = DCA1F(data, ["Residuos_ABS","Parcela"])
        h_var = aov["p-valor"]["Parcelas"] > 0.05
    elif tipo == "dbca1f":
        aov = DBCA1F(data, "Residuos_ABS")
        h_var = aov["p-valor"]["Parcelas"] > 0.05 and aov["p-valor"]["Ambientes"] > 0.05
    elif tipo == "dca2f":
        aov = DCA1F(data, ["Residuos_ABS","Tratamiento"])
        h_var = aov["p-valor"]["Tratamientos"] > 0.05

    aov.to_excel(writer, "HOMOCEDASTICIDAD" + ambiente)
    arcpy.AddMessage("Homocedasticidad")
    arcpy.AddMessage("Levene")
    arcpy.AddMessage(aov)
    
    """
    AGREGAR + ambiente
    #data.plot.scatter("Predichos", "Residuos_Est")
    #arcpy.AddMessage("Ver gráfico")
    #plt.show()
    #plt.savefig(resvspred)
    """

    arcpy.AddMessage("")

    #
    # ADITIVIDAD
    #

    """
    if tipo == "dca2f" or tipo == "dcba1f":
        arcpy.AddMessage("Aditividad")
        data.groupby(["Parcela","Ambiente"])["Rinde"].mean().unstack("Parcela").plot.line()
        arcpy.AddMessage("Ver gráfico")
        #plt.show()
        plt.savefig(aditividad)
        arcpy.AddMessage("")
    """
    
    arcpy.AddMessage("")

    return pvalue > 0.05 and h_var

def DMS(k,gl,cme,r):
    """
    Q(alfa, tratamientos (k), gl error) * rtsq(CMe/repeticiones)
    Q(5%,3,8)=4.41
    4.41 * rtsq(0.43333/5) = 1.18
    """

    q = pd.DataFrame([[17.969, 27.066, 32.925, 37.149, 40.481, 43.203, 45.501, 47.482, 49.22, 50.78, 52.161, 53.346, 54.469, 55.534, 56.49, 57.353, 58.176, 58.945, 59.666],
    [6.085, 8.344, 9.813, 10.891, 11.744, 12.444, 13.039, 13.553, 14.003, 14.407, 14.761, 15.086, 15.386, 15.662, 15.921, 16.157, 16.38, 16.588, 16.785],
    [4.501, 5.914, 6.829, 7.504, 8.039, 8.48, 8.855, 9.18, 9.465, 9.721, 9.948, 10.156, 10.347, 10.524, 10.689, 10.841, 10.985, 11.119, 11.245],
    [3.926, 5.044, 5.761, 6.29, 6.709, 7.055, 7.349, 7.605, 7.829, 8.031, 8.212, 8.376, 8.527, 8.666, 8.796, 8.918, 9.031, 9.138, 9.238],
    [3.635, 4.605, 5.221, 5.676, 6.035, 6.332, 6.585, 6.804, 6.997, 7.171, 7.325, 7.468, 7.598, 7.718, 7.83, 7.935, 8.033, 8.125, 8.211],
    [3.46, 4.342, 4.898, 5.307, 5.63, 5.897, 6.124, 6.321, 6.495, 6.651, 6.791, 6.918, 7.036, 7.145, 7.245, 7.34, 7.428, 7.511, 7.589],
    [3.344, 4.167, 4.684, 5.063, 5.361, 5.607, 5.817, 5.999, 6.16, 6.304, 6.433, 6.552, 6.66, 6.76, 6.853, 6.941, 7.022, 7.099, 7.171],
    [3.261, 4.043, 4.531, 4.888, 5.169, 5.4, 5.598, 5.769, 5.92, 6.055, 6.177, 6.288, 6.391, 6.484, 6.572, 6.654, 6.731, 6.803, 6.871],
    [3.199, 3.951, 4.417, 4.757, 5.025, 5.246, 5.433, 5.596, 5.74, 5.869, 5.985, 6.09, 6.188, 6.277, 6.36, 6.439, 6.512, 6.58, 6.645],
    [3.151, 3.879, 4.328, 4.656, 4.913, 5.126, 5.306, 5.462, 5.6, 5.723, 5.835, 5.936, 6.029, 6.115, 6.195, 6.27, 6.34, 6.406, 6.469],
    [3.113, 3.822, 4.258, 4.575, 4.824, 5.03, 5.203, 5.354, 5.488, 5.607, 5.715, 5.812, 5.902, 5.986, 6.063, 6.135, 6.203, 6.267, 6.327],
    [3.081, 3.775, 4.2, 4.509, 4.752, 4.951, 5.12, 5.266, 5.396, 5.512, 5.616, 5.711, 5.798, 5.879, 5.954, 6.024, 6.09, 6.152, 6.21],
    [3.055, 3.736, 4.152, 4.454, 4.691, 4.886, 5.05, 5.193, 5.319, 5.432, 5.534, 5.627, 5.712, 5.79, 5.863, 5.932, 5.996, 6.056, 6.113],
    [3.033, 3.703, 4.112, 4.408, 4.64, 4.83, 4.992, 5.131, 5.254, 5.365, 5.464, 5.555, 5.638, 5.715, 5.786, 5.854, 5.916, 5.974, 6.03],
    [3.014, 3.675, 4.077, 4.368, 4.596, 4.783, 4.941, 5.078, 5.199, 5.307, 5.405, 5.493, 5.575, 5.65, 5.72, 5.786, 5.847, 5.905, 5.959],
    [2.998, 3.651, 4.047, 4.334, 4.558, 4.742, 4.898, 5.032, 5.151, 5.257, 5.353, 5.44, 5.52, 5.594, 5.663, 5.727, 5.788, 5.844, 5.897],
    [2.984, 3.63, 4.021, 4.304, 4.525, 4.706, 4.859, 4.992, 5.109, 5.213, 5.307, 5.393, 5.472, 5.545, 5.613, 5.676, 5.735, 5.791, 5.844],
    [2.971, 3.611, 3.998, 4.277, 4.496, 4.675, 4.826, 4.956, 5.071, 5.174, 5.267, 5.352, 5.43, 5.502, 5.568, 5.631, 5.689, 5.744, 5.796],
    [2.96, 3.594, 3.978, 4.254, 4.47, 4.646, 4.796, 4.925, 5.038, 5.14, 5.232, 5.315, 5.392, 5.463, 5.529, 5.59, 5.648, 5.702, 5.753],
    [2.95, 3.579, 3.96, 4.233, 4.446, 4.621, 4.769, 4.897, 5.009, 5.109, 5.2, 5.282, 5.358, 5.428, 5.493, 5.554, 5.611, 5.664, 5.715],
    [2.941, 3.566, 3.943, 4.214, 4.426, 4.599, 4.745, 4.871, 4.982, 5.081, 5.171, 5.253, 5.328, 5.397, 5.461, 5.522, 5.577, 5.63, 5.68],
    [2.933, 3.554, 3.928, 4.197, 4.407, 4.578, 4.723, 4.848, 4.958, 5.056, 5.145, 5.226, 5.3, 5.369, 5.432, 5.492, 5.548, 5.6, 5.649],
    [2.925, 3.543, 3.915, 4.182, 4.389, 4.56, 4.703, 4.827, 4.936, 5.034, 5.121, 5.202, 5.275, 5.343, 5.406, 5.465, 5.52, 5.572, 5.621],
    [2.919, 3.533, 3.902, 4.168, 4.374, 4.543, 4.685, 4.808, 4.916, 5.013, 5.1, 5.179, 5.252, 5.319, 5.382, 5.44, 5.495, 5.546, 5.595],
    [2.913, 3.524, 3.891, 4.155, 4.359, 4.527, 4.668, 4.791, 4.898, 4.993, 5.08, 5.159, 5.231, 5.298, 5.36, 5.418, 5.472, 5.523, 5.571],
    [2.907, 3.515, 3.881, 4.143, 4.346, 4.513, 4.653, 4.775, 4.881, 4.976, 5.062, 5.14, 5.212, 5.278, 5.339, 5.397, 5.451, 5.501, 5.549],
    [2.902, 3.508, 3.872, 4.132, 4.334, 4.499, 4.639, 4.76, 4.865, 4.96, 5.045, 5.123, 5.194, 5.26, 5.321, 5.378, 5.431, 5.481, 5.529],
    [2.897, 3.5, 3.863, 4.121, 4.323, 4.487, 4.626, 4.746, 4.851, 4.945, 5.029, 5.107, 5.177, 5.243, 5.303, 5.36, 5.413, 5.463, 5.51],
    [2.892, 3.494, 3.854, 4.112, 4.312, 4.476, 4.614, 4.733, 4.837, 4.931, 5.015, 5.092, 5.162, 5.227, 5.287, 5.343, 5.396, 5.446, 5.492],
    [2.888, 3.488, 3.847, 4.103, 4.302, 4.465, 4.602, 4.721, 4.825, 4.918, 5.001, 5.078, 5.148, 5.212, 5.272, 5.328, 5.38, 5.43, 5.476],
    [2.884, 3.482, 3.84, 4.095, 4.293, 4.455, 4.592, 4.71, 4.814, 4.906, 4.989, 5.065, 5.134, 5.198, 5.258, 5.314, 5.366, 5.415, 5.461],
    [2.881, 3.476, 3.833, 4.087, 4.285, 4.446, 4.582, 4.7, 4.803, 4.894, 4.977, 5.053, 5.122, 5.185, 5.245, 5.3, 5.352, 5.401, 5.446],
    [2.877, 3.471, 3.827, 4.08, 4.277, 4.438, 4.573, 4.69, 4.792, 4.884, 4.966, 5.041, 5.11, 5.173, 5.232, 5.288, 5.339, 5.387, 5.433],
    [2.874, 3.467, 3.821, 4.073, 4.269, 4.429, 4.564, 4.68, 4.783, 4.874, 4.956, 5.03, 5.099, 5.162, 5.221, 5.276, 5.327, 5.375, 5.421],
    [2.871, 3.462, 3.815, 4.067, 4.262, 4.422, 4.556, 4.672, 4.774, 4.865, 4.946, 5.02, 5.088, 5.151, 5.21, 5.264, 5.315, 5.363, 5.409],
    [2.868, 3.458, 3.81, 4.061, 4.256, 4.415, 4.548, 4.664, 4.765, 4.856, 4.937, 5.011, 5.079, 5.141, 5.199, 5.254, 5.305, 5.353, 5.398],
    [2.865, 3.454, 3.805, 4.056, 4.249, 4.408, 4.541, 4.656, 4.757, 4.847, 4.928, 5.002, 5.069, 5.132, 5.19, 5.244, 5.295, 5.342, 5.387],
    [2.863, 3.45, 3.801, 4.05, 4.244, 4.401, 4.534, 4.649, 4.749, 4.839, 4.92, 4.993, 5.061, 5.123, 5.18, 5.234, 5.285, 5.332, 5.377],
    [2.86, 3.447, 3.796, 4.045, 4.238, 4.395, 4.528, 4.642, 4.742, 4.832, 4.912, 4.985, 5.052, 5.114, 5.172, 5.226, 5.276, 5.323, 5.367],
    [2.858, 3.443, 3.792, 4.04, 4.233, 4.389, 4.521, 4.635, 4.735, 4.825, 4.905, 4.978, 5.044, 5.106, 5.163, 5.217, 5.267, 5.314, 5.358],
    [2.856, 3.44, 3.788, 4.036, 4.227, 4.384, 4.516, 4.629, 4.729, 4.818, 4.898, 4.97, 5.037, 5.098, 5.156, 5.209, 5.259, 5.306, 5.35],
    [2.854, 3.437, 3.784, 4.031, 4.223, 4.379, 4.51, 4.623, 4.723, 4.811, 4.891, 4.963, 5.03, 5.091, 5.148, 5.201, 5.251, 5.298, 5.342],
    [2.852, 3.434, 3.781, 4.027, 4.218, 4.374, 4.505, 4.618, 4.717, 4.805, 4.885, 4.957, 5.023, 5.084, 5.141, 5.194, 5.244, 5.29, 5.334],
    [2.85, 3.431, 3.777, 4.023, 4.214, 4.369, 4.5, 4.612, 4.711, 4.799, 4.879, 4.95, 5.017, 5.077, 5.134, 5.187, 5.236, 5.283, 5.327],
    [2.848, 3.429, 3.774, 4.019, 4.21, 4.364, 4.495, 4.607, 4.706, 4.794, 4.873, 4.945, 5.01, 5.071, 5.128, 5.18, 5.23, 5.276, 5.319],
    [2.847, 3.426, 3.771, 4.016, 4.206, 4.36, 4.49, 4.602, 4.701, 4.788, 4.867, 4.939, 5.004, 5.065, 5.121, 5.174, 5.223, 5.269, 5.313],
    [2.845, 3.424, 3.768, 4.012, 4.202, 4.356, 4.486, 4.598, 4.696, 4.783, 4.862, 4.933, 4.999, 5.059, 5.115, 5.168, 5.217, 5.263, 5.306],
    [2.843, 3.421, 3.765, 4.009, 4.198, 4.352, 4.481, 4.593, 4.691, 4.778, 4.857, 4.929, 4.993, 5.054, 5.11, 5.162, 5.211, 5.257, 5.3],
    [2.842, 3.419, 3.762, 4.006, 4.195, 4.348, 4.477, 4.589, 4.687, 4.774, 4.852, 4.923, 4.988, 5.048, 5.104, 5.156, 5.205, 5.251, 5.294],
    [2.84, 3.417, 3.759, 4.003, 4.191, 4.344, 4.473, 4.585, 4.682, 4.769, 4.847, 4.919, 4.983, 5.043, 5.099, 5.151, 5.2, 5.245, 5.288],
    [2.839, 3.415, 3.757, 4, 4.188, 4.341, 4.47, 4.581, 4.678, 4.765, 4.843, 4.914, 4.979, 5.038, 5.094, 5.146, 5.194, 5.24, 5.283],
    [2.838, 3.413, 3.754, 3.998, 4.185, 4.338, 4.466, 4.577, 4.674, 4.761, 4.839, 4.909, 4.974, 5.034, 5.089, 5.141, 5.189, 5.235, 5.278],
    [2.836, 3.411, 3.752, 3.995, 4.182, 4.334, 4.463, 4.573, 4.67, 4.757, 4.834, 4.905, 4.97, 5.029, 5.084, 5.136, 5.184, 5.23, 5.273],
    [2.835, 3.409, 3.75, 3.992, 4.179, 4.331, 4.459, 4.57, 4.667, 4.753, 4.83, 4.901, 4.965, 5.025, 5.08, 5.132, 5.18, 5.225, 5.268],
    [2.834, 3.407, 3.748, 3.99, 4.176, 4.328, 4.456, 4.566, 4.663, 4.749, 4.827, 4.897, 4.961, 5.021, 5.076, 5.127, 5.175, 5.22, 5.263],
    [2.833, 3.406, 3.746, 3.987, 4.174, 4.325, 4.453, 4.563, 4.66, 4.746, 4.823, 4.893, 4.958, 5.017, 5.071, 5.123, 5.171, 5.216, 5.259],
    [2.832, 3.404, 3.744, 3.985, 4.171, 4.323, 4.45, 4.56, 4.656, 4.742, 4.819, 4.89, 4.954, 5.013, 5.067, 5.119, 5.167, 5.212, 5.254],
    [2.831, 3.403, 3.742, 3.983, 4.169, 4.32, 4.447, 4.557, 4.653, 4.739, 4.816, 4.886, 4.95, 5.009, 5.064, 5.115, 5.163, 5.208, 5.25],
    [2.83, 3.401, 3.74, 3.981, 4.166, 4.317, 4.444, 4.554, 4.65, 4.736, 4.813, 4.882, 4.947, 5.005, 5.06, 5.111, 5.159, 5.204, 5.246],
    [2.829, 3.4, 3.738, 3.979, 4.164, 4.315, 4.442, 4.551, 4.647, 4.732, 4.809, 4.879, 4.943, 5.002, 5.056, 5.107, 5.155, 5.2, 5.242],
    [2.828, 3.398, 3.736, 3.977, 4.162, 4.312, 4.439, 4.548, 4.644, 4.729, 4.806, 4.876, 4.94, 4.998, 5.053, 5.104, 5.151, 5.196, 5.238],
    [2.827, 3.397, 3.735, 3.975, 4.16, 4.31, 4.437, 4.546, 4.641, 4.727, 4.803, 4.873, 4.937, 4.995, 5.049, 5.1, 5.148, 5.192, 5.234],
    [2.826, 3.395, 3.733, 3.973, 4.158, 4.308, 4.434, 4.543, 4.639, 4.724, 4.8, 4.87, 4.934, 4.992, 5.046, 5.097, 5.144, 5.189, 5.231],
    [2.825, 3.394, 3.731, 3.971, 4.156, 4.306, 4.432, 4.541, 4.636, 4.721, 4.797, 4.867, 4.93, 4.989, 5.043, 5.094, 5.141, 5.186, 5.228],
    [2.824, 3.393, 3.73, 3.969, 4.154, 4.303, 4.43, 4.538, 4.634, 4.718, 4.795, 4.864, 4.928, 4.986, 5.04, 5.091, 5.138, 5.182, 5.224],
    [2.823, 3.392, 3.728, 3.967, 4.152, 4.301, 4.427, 4.536, 4.631, 4.716, 4.792, 4.861, 4.925, 4.983, 5.037, 5.088, 5.135, 5.179, 5.221],
    [2.823, 3.391, 3.727, 3.966, 4.15, 4.299, 4.425, 4.534, 4.629, 4.713, 4.789, 4.859, 4.922, 4.98, 5.034, 5.085, 5.132, 5.176, 5.218],
    [2.822, 3.389, 3.726, 3.964, 4.148, 4.298, 4.423, 4.531, 4.627, 4.711, 4.787, 4.856, 4.919, 4.978, 5.031, 5.082, 5.129, 5.173, 5.215],
    [2.821, 3.388, 3.724, 3.963, 4.146, 4.296, 4.421, 4.529, 4.624, 4.709, 4.785, 4.854, 4.917, 4.975, 5.029, 5.079, 5.126, 5.17, 5.212],
    [2.821, 3.387, 3.723, 3.961, 4.145, 4.294, 4.419, 4.527, 4.622, 4.706, 4.782, 4.851, 4.914, 4.972, 5.026, 5.076, 5.123, 5.167, 5.209],
    [2.82, 3.386, 3.722, 3.96, 4.143, 4.292, 4.417, 4.525, 4.62, 4.704, 4.78, 4.849, 4.912, 4.97, 5.024, 5.074, 5.121, 5.165, 5.206],
    [2.819, 3.385, 3.72, 3.958, 4.141, 4.29, 4.415, 4.523, 4.618, 4.702, 4.778, 4.847, 4.91, 4.968, 5.021, 5.071, 5.118, 5.162, 5.203],
    [2.818, 3.384, 3.719, 3.957, 4.14, 4.289, 4.414, 4.521, 4.616, 4.7, 4.776, 4.844, 4.907, 4.965, 5.019, 5.069, 5.115, 5.159, 5.201],
    [2.818, 3.383, 3.718, 3.955, 4.138, 4.287, 4.412, 4.52, 4.614, 4.698, 4.774, 4.842, 4.905, 4.963, 5.017, 5.066, 5.113, 5.157, 5.198],
    [2.817, 3.382, 3.717, 3.954, 4.137, 4.285, 4.41, 4.518, 4.612, 4.696, 4.771, 4.84, 4.903, 4.961, 5.014, 5.064, 5.11, 5.154, 5.196],
    [2.817, 3.381, 3.716, 3.953, 4.136, 4.284, 4.409, 4.516, 4.61, 4.694, 4.769, 4.838, 4.901, 4.958, 5.012, 5.062, 5.108, 5.152, 5.193],
    [2.816, 3.381, 3.715, 3.952, 4.134, 4.282, 4.407, 4.514, 4.609, 4.692, 4.768, 4.836, 4.899, 4.956, 5.01, 5.059, 5.106, 5.15, 5.191],
    [2.815, 3.38, 3.714, 3.95, 4.133, 4.281, 4.405, 4.513, 4.607, 4.69, 4.766, 4.834, 4.897, 4.954, 5.008, 5.057, 5.104, 5.147, 5.189],
    [2.815, 3.379, 3.713, 3.949, 4.132, 4.279, 4.404, 4.511, 4.605, 4.689, 4.764, 4.832, 4.895, 4.952, 5.006, 5.055, 5.102, 5.145, 5.186],
    [2.814, 3.378, 3.712, 3.948, 4.13, 4.278, 4.402, 4.509, 4.604, 4.687, 4.762, 4.83, 4.893, 4.95, 5.004, 5.053, 5.099, 5.143, 5.184],
    [2.814, 3.377, 3.711, 3.947, 4.129, 4.277, 4.401, 4.508, 4.602, 4.685, 4.76, 4.828, 4.891, 4.948, 5.002, 5.051, 5.097, 5.141, 5.182],
    [2.813, 3.377, 3.71, 3.946, 4.128, 4.275, 4.4, 4.506, 4.6, 4.684, 4.759, 4.827, 4.889, 4.946, 5, 5.049, 5.095, 5.139, 5.18],
    [2.813, 3.376, 3.709, 3.945, 4.127, 4.274, 4.398, 4.505, 4.599, 4.682, 4.757, 4.825, 4.887, 4.945, 4.998, 5.048, 5.093, 5.137, 5.178],
    [2.812, 3.375, 3.708, 3.944, 4.125, 4.273, 4.397, 4.504, 4.597, 4.68, 4.755, 4.823, 4.886, 4.943, 4.996, 5.046, 5.092, 5.135, 5.176],
    [2.812, 3.374, 3.707, 3.943, 4.124, 4.272, 4.395, 4.502, 4.596, 4.679, 4.754, 4.822, 4.884, 4.941, 4.994, 5.044, 5.09, 5.133, 5.174],
    [2.811, 3.374, 3.706, 3.942, 4.123, 4.27, 4.394, 4.501, 4.594, 4.677, 4.752, 4.82, 4.882, 4.939, 4.992, 5.042, 5.088, 5.131, 5.172],
    [2.811, 3.373, 3.705, 3.941, 4.122, 4.269, 4.393, 4.5, 4.593, 4.676, 4.751, 4.819, 4.88, 4.938, 4.991, 5.04, 5.086, 5.129, 5.17],
    [2.81, 3.372, 3.704, 3.94, 4.121, 4.268, 4.392, 4.498, 4.592, 4.675, 4.749, 4.817, 4.879, 4.936, 4.989, 5.039, 5.084, 5.128, 5.168],
    [2.81, 3.372, 3.704, 3.939, 4.12, 4.267, 4.391, 4.497, 4.59, 4.673, 4.748, 4.815, 4.877, 4.935, 4.987, 5.037, 5.083, 5.126, 5.167],
    [2.809, 3.371, 3.703, 3.938, 4.119, 4.266, 4.389, 4.496, 4.589, 4.672, 4.746, 4.814, 4.876, 4.933, 4.986, 5.035, 5.081, 5.124, 5.165],
    [2.809, 3.37, 3.702, 3.937, 4.118, 4.265, 4.388, 4.495, 4.588, 4.67, 4.745, 4.813, 4.874, 4.931, 4.984, 5.034, 5.08, 5.123, 5.163],
    [2.809, 3.37, 3.701, 3.936, 4.117, 4.264, 4.387, 4.493, 4.586, 4.669, 4.744, 4.811, 4.873, 4.93, 4.983, 5.032, 5.078, 5.121, 5.162],
    [2.808, 3.369, 3.701, 3.935, 4.116, 4.263, 4.386, 4.492, 4.585, 4.668, 4.742, 4.81, 4.872, 4.929, 4.981, 5.031, 5.077, 5.119, 5.16],
    [2.808, 3.369, 3.7, 3.934, 4.115, 4.262, 4.385, 4.491, 4.584, 4.667, 4.741, 4.809, 4.87, 4.927, 4.98, 5.029, 5.075, 5.118, 5.158],
    [2.807, 3.368, 3.699, 3.934, 4.115, 4.261, 4.384, 4.49, 4.583, 4.665, 4.74, 4.807, 4.869, 4.926, 4.978, 5.028, 5.074, 5.116, 5.157],
    [2.807, 3.367, 3.698, 3.933, 4.114, 4.26, 4.383, 4.489, 4.582, 4.664, 4.739, 4.806, 4.868, 4.924, 4.977, 5.026, 5.072, 5.115, 5.155],
    [2.807, 3.367, 3.698, 3.932, 4.113, 4.259, 4.382, 4.488, 4.581, 4.663, 4.737, 4.805, 4.866, 4.923, 4.976, 5.025, 5.071, 5.113, 5.154],
    [2.806, 3.366, 3.697, 3.931, 4.112, 4.258, 4.381, 4.487, 4.58, 4.662, 4.736, 4.803, 4.865, 4.922, 4.974, 5.023, 5.069, 5.112, 5.152],
    [2.806, 3.366, 3.696, 3.931, 4.111, 4.257, 4.38, 4.486, 4.579, 4.661, 4.735, 4.802, 4.864, 4.92, 4.973, 5.022, 5.068, 5.111, 5.151],
    [2.806, 3.365, 3.696, 3.93, 4.11, 4.256, 4.379, 4.485, 4.577, 4.66, 4.734, 4.801, 4.862, 4.919, 4.972, 5.021, 5.067, 5.109, 5.15],
    [2.805, 3.365, 3.695, 3.929, 4.11, 4.255, 4.378, 4.484, 4.576, 4.659, 4.733, 4.8, 4.861, 4.918, 4.97, 5.02, 5.065, 5.108, 5.148],
    [2.805, 3.364, 3.695, 3.928, 4.109, 4.255, 4.377, 4.483, 4.575, 4.658, 4.732, 4.799, 4.86, 4.917, 4.969, 5.018, 5.064, 5.107, 5.147],
    [2.805, 3.364, 3.694, 3.928, 4.108, 4.254, 4.376, 4.482, 4.574, 4.657, 4.73, 4.798, 4.859, 4.916, 4.968, 5.017, 5.063, 5.106, 5.146],
    [2.804, 3.363, 3.693, 3.927, 4.107, 4.253, 4.375, 4.481, 4.573, 4.656, 4.729, 4.797, 4.858, 4.914, 4.967, 5.016, 5.061, 5.104, 5.144],
    [2.804, 3.363, 3.693, 3.926, 4.107, 4.252, 4.375, 4.48, 4.573, 4.655, 4.728, 4.795, 4.857, 4.913, 4.966, 5.015, 5.06, 5.103, 5.143],
    [2.804, 3.362, 3.692, 3.926, 4.106, 4.251, 4.374, 4.479, 4.572, 4.654, 4.727, 4.794, 4.856, 4.912, 4.964, 5.013, 5.059, 5.102, 5.142],
    [2.803, 3.362, 3.692, 3.925, 4.105, 4.251, 4.373, 4.478, 4.571, 4.653, 4.727, 4.793, 4.855, 4.911, 4.963, 5.012, 5.058, 5.101, 5.141],
    [2.803, 3.362, 3.691, 3.925, 4.104, 4.25, 4.372, 4.477, 4.57, 4.652, 4.726, 4.792, 4.854, 4.91, 4.962, 5.011, 5.057, 5.099, 5.14],
    [2.803, 3.361, 3.691, 3.924, 4.104, 4.249, 4.371, 4.477, 4.569, 4.651, 4.725, 4.791, 4.853, 4.909, 4.961, 5.01, 5.056, 5.098, 5.138],
    [2.803, 3.361, 3.69, 3.923, 4.103, 4.248, 4.371, 4.476, 4.568, 4.65, 4.724, 4.79, 4.852, 4.908, 4.96, 5.009, 5.054, 5.097, 5.137],
    [2.802, 3.36, 3.69, 3.923, 4.102, 4.248, 4.37, 4.475, 4.567, 4.649, 4.723, 4.789, 4.851, 4.907, 4.959, 5.008, 5.053, 5.096, 5.136],
    [2.802, 3.36, 3.689, 3.922, 4.102, 4.247, 4.369, 4.474, 4.566, 4.648, 4.722, 4.789, 4.85, 4.906, 4.958, 5.007, 5.052, 5.095, 5.135],
    [2.802, 3.36, 3.689, 3.922, 4.101, 4.246, 4.368, 4.473, 4.566, 4.647, 4.721, 4.788, 4.849, 4.905, 4.957, 5.006, 5.051, 5.094, 5.134],
    [2.801, 3.359, 3.688, 3.921, 4.101, 4.246, 4.368, 4.473, 4.565, 4.646, 4.72, 4.787, 4.848, 4.904, 4.956, 5.005, 5.05, 5.093, 5.133],
    [2.801, 3.359, 3.688, 3.92, 4.1, 4.245, 4.367, 4.472, 4.564, 4.646, 4.719, 4.786, 4.847, 4.903, 4.955, 5.004, 5.049, 5.092, 5.132],
    [2.801, 3.358, 3.687, 3.92, 4.099, 4.244, 4.366, 4.471, 4.563, 4.645, 4.718, 4.785, 4.846, 4.902, 4.954, 5.003, 5.048, 5.091, 5.131],
    [2.801, 3.358, 3.687, 3.919, 4.099, 4.244, 4.365, 4.47, 4.562, 4.644, 4.718, 4.784, 4.845, 4.901, 4.953, 5.002, 5.047, 5.09, 5.13],
    [2.8, 3.358, 3.686, 3.919, 4.098, 4.243, 4.365, 4.47, 4.562, 4.643, 4.717, 4.783, 4.844, 4.9, 4.952, 5.001, 5.046, 5.089, 5.129],
    [2.8, 3.357, 3.686, 3.918, 4.098, 4.243, 4.364, 4.469, 4.561, 4.642, 4.716, 4.782, 4.843, 4.899, 4.951, 5, 5.045, 5.088, 5.128],
    [2.8, 3.357, 3.685, 3.918, 4.097, 4.242, 4.364, 4.468, 4.56, 4.642, 4.715, 4.782, 4.842, 4.899, 4.95, 4.999, 5.044, 5.087, 5.127]])

    #columns (k) = [2-20]
    #rows (gl) = [1-120]
    #alfa = 0.05
    
    if k - 1 > len(q.columns):
        k = len(q.columns) - 1
    else:
        k = k - 2

    if gl > len(q.index):
        gl = len(q.index) - 1
    else:
        gl = gl - 1

    return q.iat[gl,k] * math.sqrt(cme/r)

def _MediaComparision(nombres,medias,nombre,dms):

    m_comp = pd.DataFrame([nombres, medias]).transpose()
    m_comp.columns = [nombre, "Medias"]

    m_comp.sort_values(by=["Medias"], ascending=False, inplace=True)
    m_comp.reset_index(drop=True, inplace=True)

    letras_lst = []

    for x in range(len(m_comp.Medias)):
        letra = []

        for y in range(len(m_comp.Medias)):
            if y < x or m_comp.Medias[x] - m_comp.Medias[y] > dms:
                letra.append("")
            elif y == x or m_comp.Medias[x] - m_comp.Medias[y] <= dms:
                letra.append("x")
            
        if len(letras_lst) == 0 or max([i for i in range(len(m_comp.Medias)) if letras_lst[-1][i] == "x"]) != max([i for i in range(len(m_comp.Medias)) if letra[i] == "x"]):
            letras_lst.append(letra)

    letras_df = pd.DataFrame(letras_lst).transpose()

    return pd.concat([m_comp,letras_df], axis=1)

def MediaComparision(data,df_e,ms_e,writer,nombre_ambiente):
    """
    Tukey 5%
    """

    arcpy.AddMessage("COMPARACIÓN DE MEDIAS")
    arcpy.AddMessage("")

    parcelas = data.Parcela.unique()
    ambientes = data.Ambiente.unique()

    for parcela in parcelas:
        for ambiente in ambientes:
            r = len(data[(data.Parcela == parcela) & (data.Ambiente == ambiente)].Rinde)
            break
        break

    medias = [data[data.Parcela == i].Rinde.mean() for i in parcelas]
    k = len(parcelas)
    n = len(ambientes) * r
    if k > 1:
        dms = DMS(k,df_e,ms_e,n)
        arcpy.AddMessage(" ".join(["k:",str(k),"n:",str(n),"gl:",str(df_e),"CMe:",str(ms_e),"DMS:",str(dms)]))

        m_comp_parcelas = _MediaComparision(parcelas, medias, "Parcelas", dms)
        m_comp_parcelas.to_excel(writer, "TUKEY_Parcela" + nombre_ambiente, index=False)
        arcpy.AddMessage(m_comp_parcelas)
        arcpy.AddMessage("")

    if tipo == "dca2f" or tipo == "dcba1f":
        medias = [data[data.Ambiente == i].Rinde.mean() for i in ambientes]
        n = len(parcelas) * r
        k = len(ambientes)
        if k > 1:
            dms = DMS(k,df_e,ms_e,n)
            arcpy.AddMessage(" ".join(["k:",str(k),"n:",str(n),"gl:",str(df_e),"CMe:",str(ms_e),"DMS:",str(dms)]))

            m_comp = _MediaComparision(ambientes, medias, "Ambientes", dms)
            m_comp.to_excel(writer, "TUKEY_Ambiente", index=False)
            arcpy.AddMessage(m_comp)
            arcpy.AddMessage("")

    if tipo == "dca2f":
        tratamientos = data.Tratamiento.unique()

        medias = [data[(data.Parcela == p) & (data.Ambiente == a)].Rinde.mean() for a in ambientes for p in parcelas]
        n = r
        k = len(parcelas) * len(ambientes)
        if k > 1:
            dms = DMS(k,df_e,ms_e,n)
            arcpy.AddMessage(" ".join(["k:",str(k),"n:",str(n),"gl:",str(df_e),"CMe:",str(ms_e),"DMS:",str(dms)]))

            m_comp = _MediaComparision(tratamientos, medias, "Parcelas * Ambientes", dms)
            m_comp.to_excel(writer, "TUKEY_Interaccion", index=False)
            arcpy.AddMessage(m_comp)
            arcpy.AddMessage("")
    
    arcpy.AddMessage("")

    return m_comp_parcelas, dms

def Analysis(data,writer,ambiente=""):

    if ambiente != "":
        ambiente = "_" + ambiente

    if tipo == "dbca1f":
        data = data.groupby(["Parcela","Ambiente"], as_index=False).mean()[["Parcela","Ambiente","Rinde"]]
    elif tipo == "dca2f":
        data["Tratamiento"] = [ "_".join([str(r.Parcela), str(r.Ambiente)]) for i, r in data.iterrows() ]

    arcpy.AddMessage("CABECERA DATOS")
    arcpy.AddMessage("")
    arcpy.AddMessage(data.head())
    arcpy.AddMessage("")
    arcpy.AddMessage("")

    #
    # SUPUESTOS
    #

    assumptions = Assumptions(data,writer,ambiente)

    #
    # ANOVA
    #

    arcpy.AddMessage("ANOVA")
    arcpy.AddMessage("")

    if tipo == "dca1f":
        arcpy.AddMessage("ANOVA DCA 1F")
        aov = DCA1F(data, ["Rinde","Parcela"])
    elif tipo == "dbca1f":
        arcpy.AddMessage("ANOVA DBCA 1F")
        aov = DBCA1F(data, "Rinde")
    elif tipo == "dca2f":
        arcpy.AddMessage("ANOVA DCA 2F")
        aov = DCA2F(data, "Rinde")

    aov.to_excel(writer, "ANOVA" + ambiente)
    arcpy.AddMessage(aov)
    arcpy.AddMessage("")
    arcpy.AddMessage("")

    #
    # COMPARACION DE MEDIAS
    #

    df_e = aov.gl["Error"]
    ms_e = aov.CM["Error"]
    m_comp_parcelas, dms = MediaComparision(data,df_e,ms_e,writer,ambiente)

    return assumptions, m_comp_parcelas, dms

def aux(m_comp_parcelas,dms,nombre_analisis,nombre_campo,ambiente=""):
    testigo = m_comp_parcelas[m_comp_parcelas["Parcelas"] == nombre_testigo]
    testigo_mean = testigo["Medias"].mean()

    columnas = []
    for column in m_comp_parcelas.columns:
        if column != "Parcelas" and column != "Medias" and testigo.iloc[0][column] == "x":
            columnas.append(column)

    parcelas = m_comp_parcelas[m_comp_parcelas["Parcelas"] != nombre_testigo]

    out = pd.DataFrame(columns=["Parcelas","Diferencia_estadistica","Rinde_Parcela","Rinde_Testigo"])

    for index, row in parcelas.iterrows():
        diferencia = True
        for columna in columnas:
            if row[columna] == "x":
                diferencia = False
                break
        if diferencia:
            out.loc[len(out)] = {"Parcelas": row["Parcelas"], "Diferencia_estadistica": row["Medias"] - testigo_mean, "Rinde_Parcela": row["Medias"], "Rinde_Testigo": testigo_mean}
        else:
            out.loc[len(out)] = {"Parcelas": row["Parcelas"], "Diferencia_estadistica": 0, "Rinde_Parcela": row["Medias"], "Rinde_Testigo": testigo_mean}
    
    out["Nombre"] = nombre_analisis
    out["Campo"] = nombre_campo
    out["Ambiente"] = ambiente
    out["DMS"] = dms
    out["Diferencia"] = out["Rinde_Parcela"] - out["Rinde_Testigo"]

    return out

"""
data = pd.DataFrame({
    "Rinde": [7,9,10,8,9,10,9,9,12,10,9,12,11,12,14],
    "Ambiente": [1,1,1,2,2,2,3,3,3,4,4,4,5,5,5],
    "Parcela": [1,2,3,1,2,3,1,2,3,1,2,3,1,2,3]
})

print data
print DCA1F(data, "Rinde")
exit()

tipo = "dca2f"
data = pd.DataFrame({
    "Rinde": [20,25,22,27,21,30,45,30,35,36,31,30,40,35,30,20,21,20,20,19,25,30,29,28,30,30,29,31,30,30,32,35,30,40,30,23,25,28,30,31,24,28,24,25,30,39,42,36,42,40,41,45,40,40,35,24,25,30,26,23,28,31,26,29,32,40,45,50,45,60,42,50,40,55,45,29,30,28,27,30],
    "Ambiente": [1,1,1,1,1,2,2,2,2,2,3,3,3,3,3,4,4,4,4,4,1,1,1,1,1,2,2,2,2,2,3,3,3,3,3,4,4,4,4,4,1,1,1,1,1,2,2,2,2,2,3,3,3,3,3,4,4,4,4,4,1,1,1,1,1,2,2,2,2,2,3,3,3,3,3,4,4,4,4,4],
    "Parcela": [1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,2,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,3,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4,4]
})
"""

#data = pd.read_excel(out_data_xlsx, "DATA", engine='openpyxl')

book = openpyxl.load_workbook(out_data_xlsx)
writer = pd.ExcelWriter(out_data_xlsx, engine='openpyxl')
writer.book = book

aux_sheet = pd.DataFrame(columns=["Nombre","Parcelas","Ambiente","Rinde_Parcela","Rinde_Testigo","Diferencia","DMS","Diferencia_estadistica","Campo"])

if dca1f_por_ambiente == True or dca1f_por_ambiente == "true":
    assumptions = True
    
    for ambiente in data.Ambiente.unique():
        new_data = data[data["Ambiente"] == ambiente]
        arcpy.AddMessage("Ambiente " + ambiente)
        arcpy.AddMessage("")
        supuestos, m_comp_parcelas, dms = Analysis(new_data,writer,ambiente) 
        assumptions = supuestos == True and assumptions

        aux_sheet = pd.concat([aux_sheet, aux(m_comp_parcelas,dms,nombre_analisis,nombre_campo,ambiente)], axis=0) 

else:
    assumptions, m_comp_parcelas, dms = Analysis(data,writer)
    aux_sheet = pd.concat([aux_sheet, aux(m_comp_parcelas,dms,nombre_analisis,nombre_campo)], axis=0) 

aux_sheet = pd.concat([aux_sheet, aux_forgiveness], axis=0)
aux_sheet.to_excel(writer, "aux", index=False)

writer.save()
writer.close()

if not assumptions:
    raise Exception("No se cumplen los supuestos")
